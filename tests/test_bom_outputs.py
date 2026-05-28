from __future__ import annotations

import json
import zipfile
from typing import Any, cast

from openpyxl import load_workbook

from altium_cruncher.altium_cruncher_cmd_bom import (
    _generic_bom_payload,
    _write_bom_output,
    _write_bom_xlsx,
)
from altium_cruncher.simple_xlsx import write_xlsx_table


def _sample_bom() -> list[dict]:
    return [
        {
            "designator": "R1",
            "value": "10k",
            "footprint": "R0603",
            "library_ref": "RES",
            "description": "resistor",
            "sheet": "top.SchDoc",
            "dnp": False,
            "parameters": {"MPN": "ABC", "Tolerance": "1%"},
        },
        {
            "designator": "C1",
            "value": "100nF",
            "footprint": "C0603",
            "library_ref": "CAP",
            "description": "capacitor",
            "sheet": "top.SchDoc",
            "dnp": True,
            "parameters": {"MPN": "XYZ"},
        },
    ]


def test_generic_bom_payload_has_stable_schema(tmp_path) -> None:
    source = tmp_path / "project.PrjPcb"
    source.write_text("", encoding="utf-8")

    payload = _generic_bom_payload(_sample_bom(), source=source, variant="V1")

    assert payload["schema"] == "wn.altium_cruncher.bom.v1"
    assert payload["variant"] == "V1"
    assert payload["component_count"] == 2
    assert payload["dnp_count"] == 1
    assert payload["parameter_columns"] == ["MPN", "Tolerance"]
    assert payload["components"][0]["Designator"] == "C1"
    assert json.loads(json.dumps(payload))["schema"] == payload["schema"]


def test_write_bom_xlsx_creates_openxml_workbook(tmp_path) -> None:
    output = tmp_path / "bom.xlsx"

    _write_bom_xlsx(output, _sample_bom())

    with zipfile.ZipFile(output) as zf:
        names = set(zf.namelist())
        assert "xl/workbook.xml" in names
        assert "xl/worksheets/sheet1.xml" in names
        sheet = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")
    assert "DNP" in sheet
    assert "Tolerance" in sheet
    assert "100nF" in sheet

    worksheet = cast(Any, load_workbook(output).active)
    assert worksheet["A1"].value == "DNP"
    assert worksheet["A2"].value == "Yes"
    assert worksheet["A1"].fill.fill_type == "solid"
    assert worksheet["A1"].font.bold is True
    assert worksheet.auto_filter.ref is None


def test_xlsx_table_preserves_text_and_highlights_selected_rows(tmp_path) -> None:
    """Keep package-like values as text and apply DNP review highlighting."""
    output = tmp_path / "table.xlsx"

    write_xlsx_table(
        output,
        columns=("designators", "footprint", "dnp"),
        rows=[
            {"designators": "R1", "footprint": "0603", "dnp": "No"},
            {"designators": "R2", "footprint": "0603", "dnp": "Yes"},
        ],
        sheet_name="BOM",
        highlighted_rows=[False, True],
    )

    worksheet = cast(Any, load_workbook(output).active)
    assert worksheet["B2"].value == "0603"
    assert worksheet["B2"].number_format == "@"
    assert worksheet["A2"].fill.fgColor.rgb != "00FFF2CC"
    assert worksheet["A3"].fill.fgColor.rgb == "00FFF2CC"
    assert worksheet["A1"].fill.fill_type == "solid"
    assert worksheet["A1"].font.bold is True
    assert worksheet.auto_filter.ref is None


def test_write_bom_output_supports_grouped_json_and_jlc_csv(tmp_path) -> None:
    source = tmp_path / "project.PrjPcb"
    raw_json = tmp_path / "raw.json"
    grouped_json = tmp_path / "grouped.json"
    jlc_csv = tmp_path / "jlc.csv"
    jlc_xlsx = tmp_path / "jlc.xlsx"
    bom = _sample_bom()
    bom[0]["parameters"]["LCSC"] = "C25804"

    _write_bom_output(
        raw_json,
        bom,
        output_format="raw-json",
        source=source,
        variant=None,
    )
    _write_bom_output(
        grouped_json,
        bom,
        output_format="grouped-json",
        source=source,
        variant=None,
    )
    _write_bom_output(
        jlc_csv,
        bom,
        output_format="jlc-csv",
        source=source,
        variant=None,
    )
    _write_bom_output(
        jlc_xlsx,
        bom,
        output_format="jlc-xlsx",
        source=source,
        variant=None,
    )

    raw_payload = json.loads(raw_json.read_text(encoding="utf-8"))
    assert isinstance(raw_payload, list)
    assert raw_payload[0]["parameters"]["LCSC"] == "C25804"
    assert "canonical_fields" not in raw_payload[0]

    grouped_payload = json.loads(grouped_json.read_text(encoding="utf-8"))
    assert grouped_payload["schema"] == "wn.altium_cruncher.bom.grouped.v1"
    assert grouped_payload["component_count"] == 2

    jlc_text = jlc_csv.read_text(encoding="utf-8")
    assert "Comment,Designator,Footprint,JLCPCB Part #" in jlc_text
    assert "C25804" in jlc_text

    workbook = load_workbook(jlc_xlsx)
    worksheet = cast(Any, workbook.active)
    assert worksheet["A1"].value == "Comment"
    assert worksheet["D2"].value == "C25804"
    assert worksheet["D2"].number_format == "@"
    assert worksheet.auto_filter.ref is None
